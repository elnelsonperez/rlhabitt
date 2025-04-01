import openpyxl
import os
import sys
from pathlib import Path

def analyze_excel_file(file_path):
    # Load the workbook
    print(f"Loading workbook: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    # Display sheet names (following the [Month]. [Year] pattern)
    print("\nSheet names:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Read the specified range (C4:AD5) from the first sheet
    sheet = wb[wb.sheetnames[0]]
    print(f"\nReading range C4:AD5 from sheet: {sheet.title}")
    
    # Print content of range C4:AD5
    print("Content of range C4:AD5:")
    for row in range(4, 6):  # Rows 4 and 5
        row_values = []
        for col in range(3, 30):  # Columns C to AD (3 to 30)
            cell_value = sheet.cell(row=row, column=col).value
            row_values.append(str(cell_value) if cell_value is not None else "None")
        print(f"Row {row}: {', '.join(row_values)}")

if __name__ == "__main__":
    # Get file path from command line or use default
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Default to the example file in the project root
        file_path = str(Path(__file__).parent.parent / "example.xlsx")
    
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
        
    analyze_excel_file(file_path)
