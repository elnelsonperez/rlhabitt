import openpyxl
import os

import logging

from pathlib import Path
from datetime import date

from src.logger import get_logger

logger = get_logger(__name__)

# Color legend for the condo rental system
# the first element is the color code, the second is the theme color index if applicable
COLOR_LEGEND = {
    ("#FF6B6B",5): "Pendiente de pago",
    ("#FFC000",7): "Cliente Airbnb",
    ("#2F75B5",4): "Larga estadia",
    ("#2F75B5",8): "Larga estadia",
    ("#F408FC",6): "Apto no disponible",
    ("#00FFFF",): "Cliente referido",
    ("#757171",0): "Mantenimiento",
    ("#548235",9): "Booking",
    ("#29E817",): "Cliente VRBO"
}

class CondoRentalParser:
    """
    Parser for condo rental Excel files with Spanish month abbreviation sheet names.
    Sheet names follow the pattern "[Month]. [Year]" (e.g., "Abr. 2025", "Dic. 2025")
    """
    
    # Spanish month abbreviations mapping to numeric months
    SPANISH_MONTHS = {
        "Ene.": 1,  # Enero (January)
        "Feb.": 2,  # Febrero (February)
        "Mar.": 3,  # Marzo (March)
        "Abr.": 4,  # Abril (April)
        "May.": 5,  # Mayo (May)
        "Jun.": 6,  # Junio (June)
        "Jul.": 7,  # Julio (July)
        "Ago.": 8,  # Agosto (August)
        "Sep.": 9,  # Septiembre (September)
        "Oct.": 10, # Octubre (October)
        "Nov.": 11, # Noviembre (November)
        "Dic.": 12  # Diciembre (December)
    }
    
    # Buildings to exclude from parsing
    EXCLUDED_BUILDINGS = [
        "LIMPIEZAS EXTERNAS"
    ]
    
    def __init__(self, file_path, verbose=False):
        """
        Initialize the parser with a file path.
        
        Args:
            file_path (str): Path to the Excel file
            verbose (bool, optional): Enable verbose logging. Defaults to False.
        """
        self.file_path = file_path
        self.workbook = None
        self.verbose = verbose
        
        # Configure logging
        if verbose:
            logging.basicConfig(
                level=logging.DEBUG,
                format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
        else:
            logging.basicConfig(
                level=logging.INFO,
                format='%(levelname)s: %(message)s'
            )
    
    def load_workbook(self):
        """
        Load the Excel workbook.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Loading workbook: {self.file_path}")
            # Load with data_only=True to get values instead of formulas
            # But also load with data_only=False to get comments
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.workbook_with_comments = openpyxl.load_workbook(self.file_path, data_only=False)
            logger.debug(f"Workbook loaded successfully with {len(self.workbook.sheetnames)} sheets")
            return True
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return False
    
    def get_sheet_names(self):
        """
        Get all sheet names in the workbook.
        
        Returns:
            list: List of sheet names or empty list if workbook is not loaded
        """
        if not self.workbook:
            logger.warning("Workbook not loaded, call load_workbook() first")
            return []
        
        return self.workbook.sheetnames
    
    def parse_sheet_date(self, sheet_name):
        """
        Parse the sheet name to get the month and year.
        Expected format: "[Month]. [Year]" (e.g., "Abr. 2025")
        
        Args:
            sheet_name (str): Name of the sheet
            
        Returns:
            tuple: (month_number, year) or (None, None) if parsing fails
        """
        try:
            # Split the sheet name by space
            parts = sheet_name.split(' ')
            if len(parts) != 2:
                logger.warning(f"Sheet name '{sheet_name}' does not follow the expected format '[Month]. [Year]'")
                return None, None
            
            month_abbr = parts[0]
            year_str = parts[1]
            
            # Get the month number from the abbreviation
            if month_abbr in self.SPANISH_MONTHS:
                month_num = self.SPANISH_MONTHS[month_abbr]
            else:
                logger.warning(f"Unknown month abbreviation: {month_abbr}")
                return None, None
            
            # Parse the year
            year = int(year_str)
            
            return month_num, year
        except Exception as e:
            logger.error(f"Error parsing sheet date from '{sheet_name}': {e}")
            return None, None
    
    def read_range(self, sheet_name, start_cell, end_cell):
        """
        Read a range of cells from a specific sheet.
        
        Args:
            sheet_name (str): Name of the sheet
            start_cell (str): Starting cell (e.g., "C4")
            end_cell (str): Ending cell (e.g., "AD5")
            
        Returns:
            list: List of rows, where each row is a list of cell values
        """
        if not self.workbook:
            logger.warning("Workbook not loaded, call load_workbook() first")
            return []
        
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return []
        
        sheet = self.workbook[sheet_name]
        logger.debug(f"Reading range {start_cell}:{end_cell} from sheet: {sheet_name}")
        
        # Parse the range
        start_col_letter, start_row = self._parse_cell_reference(start_cell)
        end_col_letter, end_row = self._parse_cell_reference(end_cell)
        
        start_col = openpyxl.utils.column_index_from_string(start_col_letter)
        end_col = openpyxl.utils.column_index_from_string(end_col_letter)
        
        # Extract the data
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
            
        return data
    
    def parse_building_table(self, sheet_name, building_start_row):
        """
        Parse a complete building reservation table.
        
        Args:
            sheet_name (str): Name of the sheet
            building_start_row (int): The starting row of the building table (where building name is)
            
        Returns:
            dict: Building data including name, apartments, and reservations
        """
        if not self.workbook or not self.workbook_with_comments:
            logger.warning("Workbook not loaded, call load_workbook() first")
            return {}
        
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return {}
        
        # Get both sheet versions (with and without formulas for comments)
        sheet = self.workbook[sheet_name]
        sheet_with_comments = self.workbook_with_comments[sheet_name]
        
        # Get the building name from B[building_start_row]
        building_name = sheet.cell(row=building_start_row, column=2).value
        logger.debug(f"Parsing building table: {building_name}")
        
        # Skip excluded buildings
        if building_name:
            for excluded_building in self.EXCLUDED_BUILDINGS:
                if excluded_building in str(building_name):
                    logger.info(f"Skipping excluded building '{building_name}' at row {building_start_row}")
                    return {}
        
        # Get the number of apartments from A[building_start_row]
        num_apartments_str = sheet.cell(row=building_start_row, column=1).value
        try:
            num_apartments = int(num_apartments_str)
        except (ValueError, TypeError):
            logger.warning(f"Could not parse number of apartments: {num_apartments_str}")
            # Try to determine the number of apartments by looking for non-empty cells in column A
            # Start from building_start_row + 2 (assuming headers at building_start_row + 1)
            row = building_start_row + 2
            num_apartments = 0
            while sheet.cell(row=row, column=1).value is not None:
                num_apartments += 1
                row += 1
        
        logger.debug(f"Number of apartments: {num_apartments}")
        
        # Get the month and year from the sheet name
        month_num, year = self.parse_sheet_date(sheet_name)
        if month_num is None or year is None:
            logger.warning(f"Could not parse month and year from sheet name: {sheet_name}")
            return {}
        
        # Read the calendar headers (weekdays and dates)
        weekdays = self.read_range(sheet_name, "C4", "AD4")[0]  # First row of the range
        date_numbers = self.read_range(sheet_name, "C5", "AD5")[0]  # First row of the range
        
        # Create a list of date objects for each column
        dates = []
        for date_num in date_numbers:
            if date_num is not None:
                try:
                    # Try to parse as an integer first (e.g., 1, 2, 3)
                    if isinstance(date_num, int):
                        day = date_num
                    else:
                        # Otherwise try to parse as a string (e.g., "01", "02")
                        day = int(str(date_num).strip())
                    dates.append(date(year, month_num, day))
                except (ValueError, TypeError):
                    logger.warning(f"Could not parse date: {date_num}")
                    dates.append(None)
            else:
                dates.append(None)
        
        # Process each apartment row
        apartments = []
        start_row = building_start_row + 2  # First apartment row (after the H header)
        
        for i in range(num_apartments):
            row_num = start_row + i
            
            # Get the apartment info from column B
            apt_info = sheet.cell(row=row_num, column=2).value
            if apt_info is None:
                logger.warning(f"No apartment info found at row {row_num}, column B")
                continue
                
            # Parse apartment code and owner name
            # Special handling for "OTROS APARTAMENTOS" building
            if building_name and "OTROS APARTAMENTOS" in str(building_name):
                # For "OTROS APARTAMENTOS", the owner is the entire text and code is None
                apt_code = None
                owner_name = apt_info
                logger.debug(f"OTROS APARTAMENTOS entry: owner='{owner_name}'")
            else:
                # Standard format: "APT CODE - OWNER NAME" or similar
                # The owner name is the last part after the last hyphen
                # The apartment code is everything before the last hyphen
                if '-' in apt_info:
                    # Find the last occurrence of "-"
                    last_separator = apt_info.rfind('-')
                    
                    # Split based on the last separator
                    apt_code = apt_info[:last_separator].strip()
                    owner_name = apt_info[last_separator+1:].strip()  # +1 to skip the "-"
                else:
                    # No separator found, use the whole string as apt_code
                    apt_code = apt_info
                    owner_name = ""
                
                logger.debug(f"Parsed apartment: code='{apt_code}', owner='{owner_name}'")
                
            # Process reservations for this apartment (columns C through AD)
            reservations = []
            for j, date_val in enumerate(dates):
                if date_val is None:
                    continue
                
                col = j + 3  # Column C is index 0 + 3
                cell_value = sheet.cell(row=row_num, column=col).value
                
                # Skip empty cells
                if cell_value is None:
                    continue
                
                # Get the cell color (background color)
                cell = sheet_with_comments.cell(row=row_num, column=col)
                cell_color = self._get_cell_color(cell)
                
                # Get the cell comment if any
                comment_text = None
                if cell.comment:
                    comment_text = cell.comment.text
                
                # Create a reservation entry
                reservation = {
                    "date": date_val.isoformat(),
                    "rate": str(cell_value) if cell_value is not None else None,
                    "color": cell_color,
                    "comment": comment_text
                }
                
                reservations.append(reservation)
            
            # Create the apartment entry
            apartment = {
                "code": apt_code,
                "owner": owner_name,
                "raw_text": apt_info,  # Include the raw cell text
                "reservations": reservations
            }
            
            apartments.append(apartment)
        
        # Create the building data
        building_data = {
            "name": building_name,
            "month": month_num,
            "year": year,
            "apartments": apartments
        }
        
        return building_data
    
    def find_building_rows(self, sheet_name):
        """
        Find all building table starting rows in the sheet.
        
        Args:
            sheet_name (str): Name of the sheet
            
        Returns:
            list: List of row numbers where building tables start
        """
        if not self.workbook:
            logger.warning("Workbook not loaded, call load_workbook() first")
            return []
        
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return []
        
        sheet = self.workbook[sheet_name]
        building_rows = []
        
        # Scan through rows looking for building names in column B
        # and apartment counts in column A
        for row in range(6, 122):  # Reasonable range to check
            cell_a = sheet.cell(row=row, column=1).value
            cell_b = sheet.cell(row=row, column=2).value
            
            # If we have a value in cell B but not in cell A, and the next row has "H" in column A, 
            # it's likely a building table header
            if (cell_b is not None and 
                (isinstance(cell_a, (int, float)) or 
                 (isinstance(cell_a, str) and cell_a.isdigit()))):
                
                # Check if the next row has "H" in column A (header row)
                if sheet.cell(row=row+1, column=1).value == "H":
                    logger.debug(f"Found building at row {row}: {cell_b}")
                    building_rows.append(row)
        
        return building_rows
    
    def parse_sheet(self, sheet_name):
        """
        Parse an entire sheet, finding and processing all building tables.
        
        Args:
            sheet_name (str): Name of the sheet
            
        Returns:
            dict: Complete sheet data including all buildings
        """
        if not self.workbook:
            logger.warning("Workbook not loaded, call load_workbook() first")
            return {}
        
        # Get month and year from sheet name
        month_num, year = self.parse_sheet_date(sheet_name)
        if month_num is None or year is None:
            logger.warning(f"Could not parse month and year from sheet name: {sheet_name}")
            return {}
        
        # Find all building rows
        building_rows = self.find_building_rows(sheet_name)
        if not building_rows:
            logger.warning(f"No building tables found in sheet: {sheet_name}")
            return {}
        
        # Parse each building table
        buildings = []
        for row in building_rows:
            building_data = self.parse_building_table(sheet_name, row)
            if building_data:
                buildings.append(building_data)
        
        # Create the sheet data
        sheet_data = {
            "month": month_num,
            "year": year,
            "buildings": buildings
        }
        
        return sheet_data
    
    def _get_cell_color(self, cell):
        """
        Extract standardized color information from a cell.
        Uses the COLOR_LEGEND to convert theme colors to RGB where possible.
        
        Args:
            cell: An openpyxl cell object
            
        Returns:
            dict: Color information with type and value, or None if no fill
        """
        # Default to no color
        if not cell or cell.fill.fill_type == 'none' or cell.fill.start_color.index == '00000000':
            return None
            
        color_info = {
            "type": "unknown",
            "value": None
        }
        
        try:
            # Get the color index/value
            color_index = cell.fill.start_color.index
            
            # Handle different color types
            if isinstance(color_index, int):
                # Theme color or indexed color
                # Try to get RGB representation
                rgb_found = False
                try:
                    rgb = cell.fill.start_color.rgb
                    if rgb and isinstance(rgb, str):
                        if rgb.startswith('FF'):
                            rgb_hex = f"#{rgb[2:]}"  # Remove alpha
                        else:
                            rgb_hex = f"#{rgb}"
                            
                        # Check if this RGB value is in our COLOR_LEGEND
                        legend_match = False
                        for legend_key, legend_value in COLOR_LEGEND.items():
                            if legend_key[0].upper() == rgb_hex.upper():
                                # We found a match in the legend
                                color_info["type"] = "rgb"
                                color_info["value"] = rgb_hex
                                color_info["meaning"] = legend_value
                                legend_match = True
                                rgb_found = True
                                break
                                
                        if not legend_match:
                            # No legend match but we still have the RGB
                            color_info["type"] = "rgb"
                            color_info["value"] = rgb_hex
                            rgb_found = True
                except:
                    pass  # No RGB available
                
                # If we couldn't get RGB, check if the theme index is in our legend
                if not rgb_found:
                    for legend_key, legend_value in COLOR_LEGEND.items():
                        if len(legend_key) > 1 and legend_key[1] == color_index:
                            # Theme index matched in the legend
                            color_info["type"] = "rgb"
                            color_info["value"] = legend_key[0]
                            color_info["meaning"] = legend_value
                            rgb_found = True
                            break
                
                # If we still don't have RGB, store as theme
                if not rgb_found:
                    color_info["type"] = "theme"
                    color_info["value"] = color_index
                    
            elif isinstance(color_index, str):
                if color_index.startswith('FF'):
                    # Standard RGB color with alpha channel
                    rgb_hex = f"#{color_index[2:]}"  # Remove alpha
                    color_info["type"] = "rgb"
                    color_info["value"] = rgb_hex
                    
                    # Check if this RGB value is in our COLOR_LEGEND
                    for legend_key, legend_value in COLOR_LEGEND.items():
                        if legend_key[0].upper() == rgb_hex.upper():
                            color_info["meaning"] = legend_value
                            break
                else:
                    # Other format (indexed, etc.)
                    # Try to get RGB
                    rgb_found = False
                    try:
                        rgb = cell.fill.start_color.rgb
                        if rgb and isinstance(rgb, str):
                            if rgb.startswith('FF'):
                                rgb_hex = f"#{rgb[2:]}"
                            else:
                                rgb_hex = f"#{rgb}"
                                
                            # Check rgb against COLOR_LEGEND
                            for legend_key, legend_value in COLOR_LEGEND.items():
                                if legend_key[0].upper() == rgb_hex.upper():
                                    color_info["type"] = "rgb"
                                    color_info["value"] = rgb_hex
                                    color_info["meaning"] = legend_value
                                    rgb_found = True
                                    break
                                    
                            if not rgb_found:
                                color_info["type"] = "rgb"
                                color_info["value"] = rgb_hex
                                rgb_found = True
                    except:
                        pass
                    
                    # If we couldn't get RGB
                    if not rgb_found:
                        color_info["type"] = "other"
                        color_info["value"] = color_index
            
            return color_info
            
        except Exception as e:
            logger.debug(f"Error extracting cell color: {e}")
            return None
    
    def _parse_cell_reference(self, cell_ref):
        """
        Parse a cell reference like "C4" into column letter and row number.
        
        Args:
            cell_ref (str): Cell reference (e.g., "C4")
            
        Returns:
            tuple: (column_letter, row_number)
        """
        # Extract column letters and row number
        col_letters = ""
        row_num = ""
        
        for char in cell_ref:
            if char.isalpha():
                col_letters += char
            elif char.isdigit():
                row_num += char
        
        return col_letters, int(row_num)

def print_range_data(file_path, sheet_name=None, verbose=False):
    """
    Test function to print the C4:AD5 range from the specified sheet.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet to read. If None, uses the first sheet.
        verbose (bool, optional): Enable verbose output. Defaults to False.
    """
    parser = CondoRentalParser(file_path, verbose=verbose)
    if not parser.load_workbook():
        return
    
    sheet_names = parser.get_sheet_names()
    if not sheet_names:
        print("No sheets found in the workbook")
        return
    
    # If no sheet name provided, use the first one
    if not sheet_name:
        sheet_name = sheet_names[0]
        print(f"Using first sheet: {sheet_name}")
    
    # Read the range C4:AD5
    data = parser.read_range(sheet_name, "C4", "AD5")
    
    if not data:
        print(f"No data found in range C4:AD5 in sheet '{sheet_name}'")
        return
    
    # Print the data
    print(f"\nContent of range C4:AD5 from sheet '{sheet_name}':")
    for i, row in enumerate(data):
        formatted_row = [str(val) if val is not None else "None" for val in row]
        print(f"Row {i+4}: {', '.join(formatted_row)}")

if __name__ == "__main__":
    import sys
    
    # Get file path from command line or use default
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = str(Path(__file__).parent.parent / "example.xlsx")
    
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    # Check if a sheet name was provided
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Check if verbose flag was provided
    verbose = "-v" in sys.argv or "--verbose" in sys.argv
    
    print_range_data(file_path, sheet_name, verbose)